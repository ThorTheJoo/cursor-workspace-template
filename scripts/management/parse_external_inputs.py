#!/usr/bin/env python3
"""Parse bank OFX, manual Excel recons, and classify input sources."""

from __future__ import annotations

import csv
import re
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    from file_classifier import classify_report
except ImportError:
    classify_report = None

# Non-operational files to exclude from input inventory (templates, archives, docs)
INVENTORY_SKIP_SUFFIXES = {".md", ".pdf", ".tar.gz", ".gz", ".zip"}
INVENTORY_SKIP_NAME_PREFIXES = ("TEMPLATE_REPO", "mdd-skills", "Payment_CSV_Imports")
INVENTORY_SKIP_NAMES = {"Payment_CSV_Template.csv"}


def should_skip_inventory_file(path: Path) -> bool:
    name = path.name
    lower = name.lower()
    if name in INVENTORY_SKIP_NAMES:
        return True
    if any(name.startswith(p) for p in INVENTORY_SKIP_NAME_PREFIXES):
        return True
    if any(lower.endswith(s) for s in INVENTORY_SKIP_SUFFIXES):
        return True
    return False


def classify_file(path: Path) -> dict[str, str]:
    if classify_report:
        meta = classify_report(path)
        return {"source_type": meta["source_type"], "report": meta["report_type"]}

    name = path.name.lower()
    if path.suffix.upper() == ".TXT" and "day end" in name:
        return {"source_type": "pos_system", "report": "day_end_summary"}
    if path.suffix.upper() == ".TXT" and "month end" in name:
        return {"source_type": "pos_system", "report": "month_end_summary"}
    if path.suffix.upper() == ".TXT" and "eft" in name:
        return {"source_type": "pos_system", "report": "eft"}
    if path.suffix.upper() == ".TXT" and "cash variance" in name:
        return {"source_type": "pos_system", "report": "cash_variance"}
    if path.suffix.upper() == ".TXT" and "stock take" in name:
        return {"source_type": "pos_system", "report": "stock_take"}
    if "nett pay list" in name:
        return {"source_type": "payroll_system", "report": "nett_pay_list"}
    if path.suffix.lower() == ".ofx":
        return {"source_type": "bank_feed", "report": "bank_statement"}
    if "cash up" in name:
        return {"source_type": "manual_recon", "report": "cash_up"}
    if "schedule of accounts" in name:
        return {"source_type": "manual_recon", "report": "supplier_schedule"}
    if "deepseek" in name or "ocr" in name or "whatsapp" in name:
        return {"source_type": "ocr_whatsapp", "report": "ocr_synthesis"}
    return {"source_type": "unknown", "report": "unknown"}


def scan_input_inventory(inputs_root: Path) -> list[dict[str, Any]]:
    inventory: list[dict[str, Any]] = []
    if not inputs_root.exists():
        return inventory
    for path in sorted(inputs_root.rglob("*")):
        if not path.is_file() or path.name.startswith("."):
            continue
        if should_skip_inventory_file(path):
            continue
        meta = classify_file(path)
        inventory.append({
            "path": str(path.relative_to(inputs_root)).replace("\\", "/"),
            "name": path.name,
            "size_bytes": path.stat().st_size,
            "modified": datetime.fromtimestamp(path.stat().st_mtime).strftime("%Y-%m-%d %H:%M"),
            **meta,
        })
    return inventory


def find_ofx_file(inputs_root: Path) -> Path | None:
    """Prefer newest OFX by statement end date anywhere under the inputs root."""
    candidates = [p for p in inputs_root.rglob("*.ofx") if p.is_file() and "(1)" not in p.name]
    if not candidates:
        return None

    def end_date(path: Path) -> str:
        text = path.read_text(encoding="utf-8", errors="replace")
        m = re.search(r"<DTEND>(\d{8})", text)
        return m.group(1) if m else "00000000"

    return max(candidates, key=end_date)


def _ofx_fmt_dt(d: str) -> str:
    return f"{d[6:8]}/{d[4:6]}/{d[0:4]}" if len(d) == 8 else d


def _ofx_categorize(memo: str) -> str:
    upper = memo.upper()
    if "WAGES" in upper or "WAGE" in upper:
        return "wages"
    if "SPEEDPOINT" in upper:
        return "speedpoint"
    return "other"


def parse_ofx(path: Path) -> dict[str, Any]:
    text = path.read_text(encoding="utf-8", errors="replace")
    acct = re.search(r"<ACCTID>(\d+)", text)
    start = re.search(r"<DTSTART>(\d{8})", text)
    end = re.search(r"<DTEND>(\d{8})", text)
    ledger = re.search(r"<LEDGERBAL>.*?<BALAMT>(-?\d+\.?\d*)", text, re.S)

    credits = 0.0
    debits = 0.0
    tx_count = 0
    memos: list[str] = []
    by_category: dict[str, float] = defaultdict(float)
    speedpoint_daily: dict[str, float] = defaultdict(float)
    wage_daily: dict[str, dict[str, float | int]] = defaultdict(
        lambda: {"total": 0.0, "count": 0}
    )

    for block in re.finditer(r"<STMTTRN>(.*?)</STMTTRN>", text, re.S):
        body = block.group(1)
        amt_m = re.search(r"<TRNAMT>(-?\d+\.?\d*)", body)
        memo_m = re.search(r"<MEMO>([^\n<]+)", body)
        dt_m = re.search(r"<DTPOSTED>(\d{8})", body)
        if not amt_m:
            continue
        amt = float(amt_m.group(1))
        memo = memo_m.group(1).strip() if memo_m else ""
        posted = _ofx_fmt_dt(dt_m.group(1)) if dt_m else ""
        cat = _ofx_categorize(memo)
        tx_count += 1
        by_category[cat] += amt
        if amt >= 0:
            credits += amt
        else:
            debits += abs(amt)
        if memo:
            memos.append(memo)
        if cat == "speedpoint" and posted:
            speedpoint_daily[posted] += amt
        if cat == "wages" and posted:
            wage_daily[posted]["total"] += abs(amt)
            wage_daily[posted]["count"] += 1

    wage_runs = sorted(
        [
            {
                "date": d,
                "total": round(v["total"], 2),
                "employee_count": int(v["count"]),
            }
            for d, v in wage_daily.items()
            if v["total"] > 0
        ],
        key=lambda x: x["date"].split("/")[::-1],
        reverse=True,
    )
    speedpoint_rows = sorted(
        [{"date": d, "total": round(v, 2)} for d, v in speedpoint_daily.items()],
        key=lambda x: x["date"].split("/")[::-1],
    )

    return {
        "source_file": path.name,
        "source_path": str(path).replace("\\", "/"),
        "source_type": "bank_feed",
        "account_id": acct.group(1) if acct else "",
        "period_start": _ofx_fmt_dt(start.group(1)) if start else "",
        "period_end": _ofx_fmt_dt(end.group(1)) if end else "",
        "ledger_balance": round(float(ledger.group(1)), 2) if ledger else None,
        "ledger_balance_date": _ofx_fmt_dt(end.group(1)) if end else "",
        "transaction_count": tx_count,
        "total_credits": round(credits, 2),
        "total_debits": round(debits, 2),
        "net_movement": round(credits - debits, 2),
        "by_category": {k: round(v, 2) for k, v in sorted(by_category.items())},
        "speedpoint_daily": speedpoint_rows,
        "wage_runs": wage_runs,
        "sample_memos": memos[:8],
    }


def _ddmmyyyy_to_date(s: str) -> datetime | None:
    parts = s.split("/")
    if len(parts) != 3:
        return None
    try:
        return datetime(int(parts[2]), int(parts[1]), int(parts[0]))
    except ValueError:
        return None


def load_payment_csv_totals(payroll_dir: Path) -> list[dict[str, Any]]:
    results: list[dict[str, Any]] = []
    if not payroll_dir.exists():
        return results
    for path in sorted(payroll_dir.glob("Payment_*.csv")):
        if " (1)" in path.name or " (2)" in path.name:
            continue
        with path.open(newline="", encoding="utf-8") as f:
            rows = list(csv.reader(f))
        if len(rows) < 4:
            continue
        pay_raw = rows[1][0].strip() if rows[1] else ""
        parts = pay_raw.split("-")
        if len(parts) != 3:
            continue
        pay_date = f"{parts[0]}/{parts[1]}/{parts[2]}"
        total = 0.0
        for row in rows[4:]:
            if len(row) <= 4:
                continue
            amount = row[4].strip()
            if not amount:
                continue
            try:
                total += float(amount)
            except ValueError as exc:
                raise ValueError(f"Invalid payment amount in {path}: {amount}") from exc
        results.append({
            "file": path.name,
            "pay_date": pay_date,
            "total": round(total, 2),
        })
    return results


def _match_wage_run(pay_date: str, wage_runs: list[dict[str, Any]], window_days: int = 3) -> dict[str, Any] | None:
    base = _ddmmyyyy_to_date(pay_date)
    if not base:
        return None
    for run in wage_runs:
        run_dt = _ddmmyyyy_to_date(run["date"])
        if not run_dt:
            continue
        delta = (run_dt - base).days
        if 0 <= delta <= window_days:
            return run
    return None


def build_bank_reconciliations(
    bank: dict[str, Any],
    canonical: dict[str, Any] | None,
    payroll_dir: Path,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    if not bank:
        return rows

    wage_runs = bank.get("wage_runs", [])
    speedpoint = {r["date"]: r["total"] for r in bank.get("speedpoint_daily", [])}

    for pay in load_payment_csv_totals(payroll_dir):
        run = _match_wage_run(pay["pay_date"], wage_runs)
        if not run:
            rows.append({
                "check": f"Payroll run {pay['pay_date']}",
                "date": pay["pay_date"],
                "source_a": "payroll_system",
                "value_a": pay["total"],
                "label_a": pay["file"],
                "source_b": "bank_feed",
                "value_b": None,
                "label_b": "Bank wage debits",
                "variance": None,
                "unit": "R",
                "status": "pending",
                "note": "No matching wage debit within 3 days on bank statement",
            })
            continue
        diff = round(pay["total"] - run["total"], 2)
        rows.append({
            "check": f"Payroll run {pay['pay_date']}",
            "date": pay["pay_date"],
            "source_a": "payroll_system",
            "value_a": pay["total"],
            "label_a": pay["file"],
            "source_b": "bank_feed",
            "value_b": run["total"],
            "label_b": f"Bank debits {run['date']} ({run['employee_count']} lines)",
            "variance": diff,
            "unit": "R",
            "status": "match" if abs(diff) < 0.02 else "review",
            "note": "",
        })

    eft = (canonical or {}).get("eft_batch", {})
    if eft.get("total_amount") and eft.get("batch_date"):
        batch_date = eft["batch_date"]
        amount = eft["total_amount"]
        base = _ddmmyyyy_to_date(batch_date)
        bank_hit: dict[str, Any] | None = None
        if base:
            for offset in range(0, 4):
                d = (base + timedelta(days=offset)).strftime("%d/%m/%Y")
                if d in speedpoint and abs(speedpoint[d] - amount) < 0.02:
                    bank_hit = {"date": d, "total": speedpoint[d]}
                    break
        if bank_hit:
            diff = round(amount - bank_hit["total"], 2)
            rows.append({
                "check": "POS EFT batch vs bank Speedpoint",
                "date": batch_date,
                "source_a": "pos_system",
                "value_a": amount,
                "label_a": f"EFT batch {batch_date} ({eft.get('total_trx', 0)} trx)",
                "source_b": "bank_feed",
                "value_b": bank_hit["total"],
                "label_b": f"Speedpoint credit {bank_hit['date']}",
                "variance": diff,
                "unit": "R",
                "status": "match" if abs(diff) < 0.02 else "review",
                "note": "Settlement typically posts T+1",
            })
        else:
            rows.append({
                "check": "POS EFT batch vs bank Speedpoint",
                "date": batch_date,
                "source_a": "pos_system",
                "value_a": amount,
                "label_a": f"EFT batch {batch_date}",
                "source_b": "bank_feed",
                "value_b": None,
                "label_b": "Speedpoint daily credits",
                "variance": None,
                "unit": "R",
                "status": "pending",
                "note": "No matching Speedpoint credit within 3 days of batch date",
            })

    return rows


def parse_schedule_of_accounts(path: Path) -> dict[str, Any]:
    if not openpyxl:
        return {"source_file": path.name, "error": "openpyxl not installed"}
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    sh = wb.active
    rows: list[dict] = []
    by_dept: dict[str, float] = {}
    total_incl = 0.0
    for row in sh.iter_rows(min_row=7, values_only=True):
        if not row or not row[3]:
            continue
        try:
            incl = float(row[8] or 0)
        except (TypeError, ValueError):
            continue
        dept = str(row[5] or "UNKNOWN").strip()
        by_dept[dept] = by_dept.get(dept, 0) + incl
        total_incl += incl
        rows.append({
            "date": str(row[1])[:10] if row[1] else "",
            "supplier": str(row[3]).strip(),
            "department": dept,
            "amount_incl": round(incl, 2),
            "paid": str(row[9] or "").strip(),
        })
    wb.close()
    top_depts = sorted(by_dept.items(), key=lambda x: x[1], reverse=True)[:6]
    return {
        "source_file": path.name,
        "source_type": "manual_recon",
        "invoice_count": len(rows),
        "total_incl": round(total_incl, 2),
        "by_department": {k: round(v, 2) for k, v in top_depts},
        "recent": rows[:5],
    }


def parse_cash_up(path: Path) -> dict[str, Any]:
    if not openpyxl:
        return {"source_file": path.name, "error": "openpyxl not installed"}
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    sheets = wb.sheetnames
    summary: dict[str, Any] = {
        "source_file": path.name,
        "source_type": "manual_recon",
        "sheet_count": len(sheets),
        "sheet_names": sheets[:12],
    }
    if sheets:
        sh = wb[sheets[0]]
        title = None
        for row in sh.iter_rows(max_row=5, values_only=True):
            if row and row[0] and "ENGEN" in str(row[0]).upper():
                title = str(row[0])
                break
        summary["site_title"] = title
    wb.close()
    return summary


def build_external_context(
    inputs_root: Path,
    canonical: dict[str, Any] | None = None,
    payroll_dir: Path | None = None,
) -> dict[str, Any]:
    ctx: dict[str, Any] = {
        "input_inventory": scan_input_inventory(inputs_root),
    }

    ofx_path = find_ofx_file(inputs_root)
    if ofx_path:
        ctx["bank_statement"] = parse_ofx(ofx_path)

    sched = [p for p in inputs_root.rglob("Schedule of Accounts*.xlsx") if "(1)" not in p.name]
    if sched:
        ctx["supplier_schedule"] = parse_schedule_of_accounts(max(sched, key=lambda p: p.stat().st_mtime))

    cash_up = [p for p in inputs_root.rglob("CASH UP*.xlsx") if "(1)" not in p.name]
    if cash_up:
        ctx["cash_up"] = parse_cash_up(max(cash_up, key=lambda p: p.stat().st_mtime))

    payroll_files = sorted(p for p in inputs_root.rglob("Nett Pay List*.xls*") if "(1)" not in p.name)
    ctx["payroll_files"] = [
        {"file": p.name, **classify_file(p), "modified": datetime.fromtimestamp(p.stat().st_mtime).isoformat()[:10]}
        for p in payroll_files
    ]

    try:
        from parse_ocr_whatsapp import build_ocr_context
        ctx.update(build_ocr_context(inputs_root, canonical))
    except ImportError as exc:
        ctx["ocr_error"] = f"OCR parser unavailable: {exc}"

    bank = ctx.get("bank_statement", {})
    if bank and payroll_dir:
        bank_rows = build_bank_reconciliations(bank, canonical, payroll_dir)
        existing = ctx.get("reconciliations", [])
        ctx["reconciliations"] = existing + bank_rows

    return ctx
